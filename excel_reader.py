import datetime
from dataclasses import dataclass
from openpyxl import load_workbook

@dataclass
class Participant:
    name: str
    lichess_link: str

@dataclass
class League:
    name: str
    participants: list[Participant]

def extract_leagues(filepath: str) -> list[League]:
    workbook = load_workbook(filename=filepath, data_only=True)
    # print(workbook.sheetnames)
    leagues_sheet = workbook["Leagues"]
    leagues_: list[League] = []
    skip: bool = False
    for col in leagues_sheet.iter_cols(): #values_only=True):
        if skip:
            skip = False
            continue
        if any(map(lambda x: x is not None, col)):
            for idx, item in enumerate(col):
                if item.value is None:
                    continue

                print(f"{col[0].column_letter}")
                leagues_.append(League(item.value, [Participant(x.value, leagues_sheet.cell(row=x.row, column=x.column+1).value) for x in col[idx+1:] if x.value is not None]))
                skip = True
                break

    return leagues_

if __name__ == "__main__":
    leagues = extract_leagues("./excel_test.xlsx")

    for league in leagues:
        print(f"{league.name}:\n\t {league.participants}")
